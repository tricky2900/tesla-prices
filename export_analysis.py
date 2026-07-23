"""
Reads the SQLite price history and exports tesla_price_analysis.xlsx with:

  1. Latest Prices     — most recent snapshot per country/model/trim
  2. Monthly Averages  — avg price per country+model, months as columns
  3. MoM Change        — live Excel formulas referencing Monthly Averages,
                         so the % change recalculates inside Excel
  4. Raw Data          — full daily history (for pivot tables / Power Query)

Point your quarterly Excel model at this file with Power Query
(Data -> Get Data -> From Workbook) and hit Refresh after each export.

Run:  python export_analysis.py
"""

import sqlite3

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from config import DB_PATH

OUT_FILE = "tesla_price_analysis.xlsx"

ARIAL = Font(name="Arial", size=10)
ARIAL_BOLD = Font(name="Arial", size=10, bold=True)


def style_sheet(ws, header_rows=1, num_fmt="#,##0", num_from_col=None):
    for row in ws.iter_rows():
        for cell in row:
            cell.font = ARIAL_BOLD if cell.row <= header_rows else ARIAL
            if (num_from_col and cell.row > header_rows
                    and cell.column >= num_from_col
                    and isinstance(cell.value, (int, float))):
                cell.number_format = num_fmt
    for col_cells in ws.columns:
        width = max((len(str(c.value)) for c in col_cells if c.value is not None),
                    default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(width + 2, 40)
    ws.freeze_panes = ws.cell(row=header_rows + 1, column=2)


def write_df(ws, df):
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)


def main():
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT * FROM prices", conn, parse_dates=["pull_date"])
    conn.close()
    if df.empty:
        raise SystemExit("Database is empty — run scraper.py first.")

    df["month"] = df["pull_date"].dt.to_period("M").astype(str)

    wb = Workbook()

    # --- 1. Latest Prices -------------------------------------------------
    latest_date = df["pull_date"].max()
    latest = (df[df["pull_date"] == latest_date]
              .loc[:, ["country", "model", "trim_name", "trim_code", "price", "currency"]]
              .sort_values(["country", "model", "price"]))
    ws = wb.active
    ws.title = "Latest Prices"
    ws.append([f"Snapshot date: {latest_date.date()}  (source: tesla.com configurator pages)"])
    write_df(ws, latest)
    style_sheet(ws, header_rows=2, num_from_col=5)

    # --- 2. Monthly Averages (country+model rows, months as columns) ------
    monthly = (df.groupby(["country", "model", "currency", "month"], as_index=False)
                 ["price"].mean())
    pivot = monthly.pivot_table(index=["country", "model", "currency"],
                                columns="month", values="price").reset_index()
    pivot.columns.name = None
    month_cols = [c for c in pivot.columns if c not in ("country", "model", "currency")]

    ws2 = wb.create_sheet("Monthly Averages")
    write_df(ws2, pivot)
    style_sheet(ws2, num_from_col=4)

    # --- 3. MoM % Change — Excel formulas against Monthly Averages --------
    ws3 = wb.create_sheet("MoM Change")
    ws3.append(["country", "model", "currency"] + month_cols[1:])
    n_rows = len(pivot)
    for i in range(n_rows):
        excel_row = i + 2          # data starts on row 2 in both sheets
        row_vals = [pivot.iloc[i]["country"], pivot.iloc[i]["model"],
                    pivot.iloc[i]["currency"]]
        for j in range(1, len(month_cols)):
            prev_col = get_column_letter(4 + j - 1)   # months start at col D
            curr_col = get_column_letter(4 + j)
            row_vals.append(
                f"=IFERROR('Monthly Averages'!{curr_col}{excel_row}"
                f"/'Monthly Averages'!{prev_col}{excel_row}-1,\"\")"
            )
        ws3.append(row_vals)
    for row in ws3.iter_rows(min_row=2, min_col=4):
        for cell in row:
            cell.number_format = "0.0%"
    style_sheet(ws3)

    # --- 4. Raw Data ------------------------------------------------------
    ws4 = wb.create_sheet("Raw Data")
    raw = df.loc[:, ["pull_date", "country", "model", "trim_code",
                     "trim_name", "price", "currency"]].copy()
    raw["pull_date"] = raw["pull_date"].dt.strftime("%Y-%m-%d")
    write_df(ws4, raw)
    style_sheet(ws4, num_from_col=6)

    wb.save(OUT_FILE)
    print(f"Wrote {OUT_FILE}: {len(latest)} latest rows, "
          f"{n_rows} country-model series across {len(month_cols)} months.")


if __name__ == "__main__":
    main()
